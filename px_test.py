#!/usr/bin/env python
# coding: utf-8

# # 第2章　最小二乗法：機械学習理論の第一歩

# ## 2.1 多項式近似と最小二乗法による推定

# **[02SE-01]**
#
# 必要なモジュールをインポートします。

# In[1]:


import numpy as np
import matplotlib.pyplot as plt
from pandas import Series, DataFrame
#Seriesは一次元、Dataframeは二次元のデータを収納するために使われる

from numpy.random import normal
#正規分布乱数
import xlwings as xw
import openpyxl as xl

#wb = xl.load_workbook("px_test.xlsm", data_only="True")
wb = xl.load_workbook("px_test.xlsm")
ws = wb["Sheet1"]


# **[02SE-02]**
#
# 正弦関数 $y=\sin(2\pi x)$ に、平均 0、標準偏差 0.3 の正規分布のノイズを載せたデータセットを生成する関数を定義します。
#
# これは、$0\le x\le 1$ の区間を等分した `num` 個の点 $\{x_n\}_{n=1}^N$ に対して、対応する $\{t_n\}_{n=1}^N$ の値を生成します。

# In[2]:


def create_dataset(num):
  xs = np.linspace(0, 1, num)
  ts = np.linspace(0, 1, num)
  i = 0
  for i in range(num):
    xs[i] = ws["A{}".format(i+1)].value
    ts[i] = ws["B{}".format(i+1)].value
  return xs, ts


# **[02SE-03]**
#
# 例として、10 個のデータをトレーニングセットとして生成します。

# In[3]:


N = 10 # データ数
xs, ts = create_dataset(N)

DataFrame({'x': xs, 't': ts})
#pandas.DataFrameは各列に異なる型のデータを格納できる。
#'ここの名前で表示': 変数


# **[02SE-04]**
#
# グラフ上にプロットすると次のようになります。

# In[4]:


fig = plt.figure(figsize=(6, 4))     #figure()は図。グラフ描画の台紙と思えばいい。figuresizeは台紙の縦横比を指定。
subplot = fig.add_subplot(1, 1, 1)    #(1, 1, 1)とはfigure()の領域内での一行目一列目の一番目という意味。上書きしたらグラフが重なったりする。
subplot.tick_params(axis='x', labelsize=12)     #params()は各軸の設定。 labelsizeは目盛りの文字サイズ
subplot.tick_params(axis='y', labelsize=12)
subplot.set_xlim(-0.05, 1.05)    #それぞれxy軸の上下限を設定。
subplot.set_ylim(-1.5, 1.5)
_ = subplot.scatter(xs, ts, marker='o', color='blue')    #散布図を描画する関数 scatter(x軸配列, y軸配列) あとは任意で追加。markerは形状を指定。
#pythonでは _(アンダースコア) も変数名として使用できる。
plt.savefig("test.png")

# **[02SE-05]**
#
# このデータに対して、最小二乗法でフィッティングした $M$ 次多項式を決定する関数を定義します。
#
# 引数 `xs`、`ts`にトレーニングセットのデータ、`m` に多項式の次数 $M$ を代入すると、多項式に対応する関数 $f(x)$ のオブジェクト、および、係数 $\{w_m\}_{m=1}^M$ の値を格納した array オブジェクトが返ります。
#
# 多項式の係数を求める際は、本文で説明した次の関係式を利用しています。
#
# $$
# \mathbf w = \left(\boldsymbol\Phi^{\rm T}\boldsymbol\Phi\right)^{-1}\boldsymbol\Phi^{\rm T}\mathbf t
# $$
#

# In[5]:


def resolve(xs, ts, m):
  phi = np.array([[x**k for k in range(m+1)] for x in xs])     #xのk乗の(m+1)×(xs)行列。 (m+1)側がk、xs側がxでループを回している
  tmp = np.linalg.inv(np.dot(phi.T, phi))      #linalg.inv()が逆行列。dot()が行列の積。つまり、phiの転置とphiの積の逆行列。
  ws = np.dot(np.dot(tmp, phi.T), ts)          #(tmpとphiの転置の積)と(ts)の積
                                               #つまりphi=Φ、tmp=逆行列項、ws=w

  def f(x):                                    #関数のネスト
    y = 0
    for i, w in enumerate(ws):                 #enumerate()はforを回しつつインデックス番号を取得出来るらしい。(f()はΣとに番号いるか..?)
      y += w * x**i                            #(2.16)のΣの式。つまり今回導出しようとしているM次多項式
    return y

  return f, ws


# **[02SE-06]**
#
# 得られた関数 $f(x)$ に対して、トレーニングセットに対する平方根平均二乗誤差を求める関数を定義します。
#
# ここでは、次式で定義される二乗誤差 $E_D$ を変数 `err` に保存した後に、平方根平均二乗誤差 $E_{\rm{RMS}}$ に変換しています。
#
# $$
# E_D=\frac{1}{2}\sum_{n=1}^N\left\{f(x_n)-t_n\right\}^2,\
# E_{\rm{RMS}} = \sqrt{\frac{2E_D}{N}}
# $$

# In[6]:


def rms_error(xs, ts, f):
  err = 0.5 * np.sum((f(xs) - ts)**2)           #二乗誤差
  return np.sqrt(2 * err / len(xs))             #平方平均二乗誤差


# **[02SE-07]**
#
# これらを用いて、結果をグラフに可視化する関数を定義します。

# In[7]:


def show_result(subplot, xs, ts, m):
  f, _ = resolve(xs, ts, m)                                      #近似多項式f()と、多項式の係数wを返す内部定義関数
  subplot.tick_params(axis='x', labelsize=12)                    #グラフの軸の設定。軸のフォントサイズ設定
  subplot.tick_params(axis='y', labelsize=12)
  subplot.set_xlim(-0.05, 1.05)                                  #グラフの上下限設定
  subplot.set_ylim(-1.5, 1.5)
  subplot.set_title('M={}'.format(m), fontsize=14)               #グラフのタイトル設定。format()は書式化の関数。

  # トレーニングセットを表示
  subplot.scatter(xs, ts, marker='o', color='blue', label=None) #scatter()散布図の描画

  linex = np.linspace(0, 1, 100)
  liney = f(linex)
  label = 'E(RMS)={:.2f}'.format(rms_error(xs, ts, f))          #平方平均二乗誤差を小数点以下二桁までラベルで表示
  subplot.plot(linex, liney, color='red', label=label)          #f(x)を赤線で描画
  subplot.legend(loc=1, fontsize=14)                            #凡例の表示(-- E(RMS)のところ)


# **[02SE-08]**
#
# 先ほど生成したトレーニングセットを用いて、0, 1, 3, 9次多項式でフィッティングした結果を表示します。

# In[8]:


fig = plt.figure(figsize=(12, 8.5))                    #グラフの縦横比
fig.subplots_adjust(wspace=0.3, hspace=0.3)            #グラフ同士の間隔
for i, m in enumerate([1, 2, 3, 4]):                   #0, 1, 3, 9でforを回す
  subplot = fig.add_subplot(2, 2, i+1)                 #二行二列で座標を導入
  show_result(subplot, xs, ts, m)                      #正解ラベル、sin2π、近似多項式を描画する内部定義関数
plt.savefig("test_result.png")

# **[02SE-09]**
#
# 最小二乗法で得られた係数 $\{w_m\}_{m=0}^M$ の具体的な値を表示します。
#
# 多項式の次数が上がると、係数の値が極端に大きくなることがわかります。

# # In[9]:
#
#
# weights = {}                               #dict型(辞書型)の生成。キー値と値をペアで保持する。
# #print(type(weights))
# for m in [0, 1, 3, 9]:
#   _, ws = resolve(xs, ts, m)               #近似多項式f()と、多項式の係数wを返す内部定義関数
#   weights['M={}'.format(m)] = Series(ws)   #{}置換フィールドに文字列型にしたmを投げる。さらに多項式の係数wsを一次元のデータ列として保持
#
# DataFrame(weights)                         #表示
#
#
# # ## 2.2 オーバーフィッティングの検出
#
# # **[02SE-10]**
# #
# # 多項式の次数を 0〜9 に変化させながら、平方根平均二乗誤差のグラフを描く関数を用意します。
#
# # In[10]:
#
#
# def show_rms_trend(xs, ts, val_xs, val_ts):
#   rms_errors = {'Training set': [], 'Test set': []}               #dict型で配列を保持??
#   for m in range(0, 10):
#     f, _ = resolve(xs, ts, m)                                     #近似多項式f()と、多項式の係数wを返す内部定義関数
#     rms_errors['Training set'].append(rms_error(xs, ts, f))       #キー"Training set"に対応する値(配列)として平方平均二乗誤差を結合
#     rms_errors['Test set'].append(rms_error(val_xs, val_ts, f))   #"Test set"に結合。しかし引数が異なる
#   df = DataFrame(rms_errors)                                      #二つの誤差をDataFrameに格納
#   ax = df.plot(figsize=(8, 5), style=['-', '--'], grid=True,     #DataFrameをプロット
#                xticks=range(0, 10),  ylim=(0, 0.9), fontsize=12)
#   ax.set_title(label='RMS Error', fontdict={'fontsize':14})
#   ax.legend(loc=0, fontsize=14)                                   #凡例
#
#
# # **[02SE-11]**
# #
# # トレーニングセットとは独立に生成したテストセットを用意します。
#
# # In[11]:
#
#
# #新たな変数に新たなデータセットを生成。
# val_xs, val_ts = create_dataset(N)
#
# DataFrame({'val_x': val_xs, 'val_t': val_ts})
#
#
# # **[02SE-12]**
# #
# # トレーニングセットとテストセットに対する平方根平均二条誤差を計算して、結果をグラフ表示にします。
# #
# # 次数が3を超えるとテストセットに対する誤差が減少しなくなることがわかります。
#
# # In[12]:
#
#
# show_rms_trend(xs, ts, val_xs, val_ts)
# #Training setで得た近似多項式を新たに生成したTest setに対して適用する事でN=3以上ではオーバーフィッティングが起こり、
# #一般性が失われていることが分かる
#
#
# # **[02SE-13]**
# #
# # 同じ計算をデータ数を増やして実施してみます。
# #
# # N = 100 でトレーニングセットとテストセットを用意します。
#
# # In[13]:
#
#
# N = 100 # データ数
# xs, ts = create_dataset(N)
# val_xs, val_ts = create_dataset(N)
#
#
# # **[02SE-14]**
# #
# # 最小二乗法でフィッティングした結果を表示します。
# #
# # 多項式の次数が上がってもオーバーフィッティングが発生しにくくなっていることがわかります。
#
# # In[14]:
#
#
# fig = plt.figure(figsize=(12, 8.5))
# fig.subplots_adjust(wspace=0.3, hspace=0.3)
# for i, m in enumerate([0, 1, 3, 9]):
#   subplot = fig.add_subplot(2, 2, i+1)
#   show_result(subplot, xs, ts, m)
#
#
# # **[02SE-15]**
# #
# # トレーニングセットとテストセットに対する平方根平均二乗誤差の変化を表示します。
# #
# # 次数が 3 を超えると平方根平均二乗誤差が約 0.3 で一定になります。これは、このデータが本質的に ±0.3 程度の誤差を持っている事を示します。
#
# # In[15]:
#
#
# show_rms_trend(xs, ts, val_xs, val_ts)
#
#
# # In[ ]:
